from __future__ import annotations

import argparse
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass
import json
import os
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

try:
    from common import load_config
except ModuleNotFoundError:  # pragma: no cover - import fallback for tests
    from .common import load_config

try:
    from generate_outline_deepseek import outline_complete
except ModuleNotFoundError:  # pragma: no cover - import fallback for tests
    from .generate_outline_deepseek import outline_complete


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path("D:/video/学科+播放页地址.xlsx")
DEFAULT_OUTPUT_ROOT = Path("D:/video/output")
INVALID_NAME_CHARS_RE = re.compile(r'[\\/:*?"<>|]')
PART_SECONDS_CHOICES = {0, 900, 1800, 3600}


@dataclass(frozen=True)
class BatchTask:
    row_number: int
    seq: int
    title: str
    page_url: str
    output_dir: Path
    progress_key: str


@dataclass(frozen=True)
class TaskResult:
    progress_key: str
    output_dir: str
    outline: str
    returncode: int
    stdout: str
    stderr: str


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def find_column(headers: list[object], needle: str) -> int:
    for index, header in enumerate(headers):
        if needle in normalize_header(header):
            return index
    raise RuntimeError(f"Missing xlsx column containing {needle!r}")


def sanitize_title(value: object) -> str:
    title = str(value or "").strip()
    title = INVALID_NAME_CHARS_RE.sub("_", title)
    title = re.sub(r"\s+", " ", title).strip()
    return title[:50] or "未命名课程"


def parse_seq(value: object) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError("Missing sequence number")
    return int(float(str(value).strip()))


def read_tasks(xlsx_path: Path, output_root: Path) -> list[BatchTask]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    try:
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        seq_col = find_column(headers, "序号")
        title_col = find_column(headers, "课程名称")
        url_col = find_column(headers, "视频链接")

        tasks: list[BatchTask] = []
        for row_number, row in enumerate(rows, start=2):
            values = list(row)
            seq_value = values[seq_col] if seq_col < len(values) else None
            title_value = values[title_col] if title_col < len(values) else None
            url_value = values[url_col] if url_col < len(values) else None
            if not any(str(item or "").strip() for item in (seq_value, title_value, url_value)):
                continue
            if not str(url_value or "").strip():
                continue

            seq = parse_seq(seq_value)
            title = str(title_value or "").strip()
            safe_title = sanitize_title(title)
            dirname = f"{seq:03d}_{safe_title}"
            tasks.append(
                BatchTask(
                    row_number=row_number,
                    seq=seq,
                    title=title or safe_title,
                    page_url=str(url_value).strip(),
                    output_dir=output_root / dirname,
                    progress_key=dirname,
                )
            )
        return tasks
    finally:
        workbook.close()


def filter_tasks(
    tasks: list[BatchTask],
    start_row: int | None,
    end_row: int | None,
) -> list[BatchTask]:
    filtered: list[BatchTask] = []
    for task in tasks:
        if start_row is not None and task.row_number < start_row:
            continue
        if end_row is not None and task.row_number > end_row:
            continue
        filtered.append(task)
    return filtered


def load_progress(progress_path: Path) -> dict[str, dict[str, str]]:
    if not progress_path.exists():
        return {}
    payload = json.loads(progress_path.read_text(encoding="utf-8-sig"))
    if not isinstance(payload, dict):
        raise RuntimeError(f"Invalid batch progress file: {progress_path}")
    progress: dict[str, dict[str, str]] = {}
    for key, value in payload.items():
        progress[str(key)] = value if isinstance(value, dict) else {"status": str(value)}
    return progress


def save_progress(progress_path: Path, progress: dict[str, dict[str, str]]) -> None:
    progress_path.parent.mkdir(parents=True, exist_ok=True)
    progress_path.write_text(
        json.dumps(progress, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def mark_progress(
    progress_path: Path,
    progress: dict[str, dict[str, str]],
    task: BatchTask,
    status: str,
    *,
    outline: str | None = None,
    error: str | None = None,
) -> None:
    item: dict[str, str] = {"status": status}
    if outline:
        item["outline"] = outline
    if error:
        item["error"] = error
    progress[task.progress_key] = item
    save_progress(progress_path, progress)


def should_skip_task(
    task: BatchTask,
    progress: dict[str, dict[str, str]],
    resume: bool,
) -> bool:
    if not resume:
        return False
    if outline_complete(task.output_dir):
        return True
    return False


def skipped_outline_path(task: BatchTask, resume: bool) -> str | None:
    outline_path = task.output_dir / "outline.md"
    if resume and outline_complete(task.output_dir):
        return str(outline_path)
    return None


def build_pipeline_command(
    task: BatchTask,
    part_seconds: int,
    resume: bool,
    *,
    skip_clean: bool = False,
    skip_context_infer: bool = False,
) -> list[str]:
    cmd = [
        sys.executable,
        str(ROOT / "scripts" / "run_pipeline.py"),
        "--page-url",
        task.page_url,
        "--title",
        task.title,
        "--seq",
        str(task.seq),
        "--output-dir",
        str(task.output_dir),
        "--part-seconds",
        str(part_seconds),
    ]
    if resume:
        cmd.append("--resume")
    if skip_clean:
        cmd.append("--skip-clean")
    if skip_context_infer:
        cmd.append("--skip-context-infer")
    return cmd


def validate_worker_policy(
    *,
    workers: int,
    skip_context_infer: bool,
    skip_clean: bool,
) -> None:
    llm_active = True
    if workers > 1 and llm_active:
        raise ValueError(
            "--workers > 1 is disabled while LLM steps are active. "
            "Use --workers 1 for context, clean, and outline generation."
        )


def run_task(
    task: BatchTask,
    part_seconds: int,
    resume: bool,
    *,
    skip_clean: bool = False,
    skip_context_infer: bool = False,
) -> TaskResult:
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    completed = subprocess.run(
        build_pipeline_command(
            task,
            part_seconds,
            resume,
            skip_clean=skip_clean,
            skip_context_infer=skip_context_infer,
        ),
        check=False,
        cwd=str(ROOT),
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return TaskResult(
        progress_key=task.progress_key,
        output_dir=str(task.output_dir),
        outline=str(task.output_dir / "outline.md"),
        returncode=completed.returncode,
        stdout=completed.stdout,
        stderr=completed.stderr,
    )


def print_dry_run(tasks: list[BatchTask]) -> None:
    for task in tasks:
        print(f"[{task.seq:03d}] {task.title}  ->  {task.output_dir}/")
        print(f"      page_url: {task.page_url}")
    print(f"共 {len(tasks)} 条")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run pipelines for videos listed in xlsx.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--workers", type=int, default=1, help="Parallel videos, 1-5.")
    parser.add_argument(
        "--part-seconds",
        type=int,
        default=900,
        choices=sorted(PART_SECONDS_CHOICES),
        help="Audio part length: 0, 900, 1800, or 3600 seconds.",
    )
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--resume", action="store_true", help="Skip entries with outline.md.")
    parser.add_argument("--skip-clean", action="store_true", help="Forward --skip-clean to run_pipeline.py.")
    parser.add_argument(
        "--skip-context-infer",
        action="store_true",
        help="Forward --skip-context-infer to run_pipeline.py.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Print tasks without executing.")
    parser.add_argument(
        "--start-row",
        type=int,
        help="1-based xlsx row number. Row 2 is the first data row when row 1 is the header.",
    )
    parser.add_argument("--end-row", type=int, help="Inclusive 1-based xlsx row number.")
    return parser.parse_args()


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")

    args = parse_args()
    load_config()

    if args.workers < 1 or args.workers > 5:
        raise ValueError("--workers must be between 1 and 5.")
    validate_worker_policy(
        workers=args.workers,
        skip_context_infer=args.skip_context_infer,
        skip_clean=args.skip_clean,
    )
    if args.start_row is not None and args.start_row < 1:
        raise ValueError("--start-row must be a positive row number.")
    if args.end_row is not None and args.end_row < 1:
        raise ValueError("--end-row must be a positive row number.")
    if (
        args.start_row is not None
        and args.end_row is not None
        and args.start_row > args.end_row
    ):
        raise ValueError("--start-row cannot be greater than --end-row.")

    tasks = filter_tasks(read_tasks(args.xlsx, args.output_root), args.start_row, args.end_row)
    if args.dry_run:
        print_dry_run(tasks)
        return

    progress_path = args.output_root / "batch_progress.json"
    progress = load_progress(progress_path)
    for task in tasks:
        if task.progress_key not in progress:
            mark_progress(progress_path, progress, task, "pending")

    runnable = [
        task
        for task in tasks
        if not should_skip_task(task, progress, args.resume)
    ]
    skipped = len(tasks) - len(runnable)
    for task in tasks:
        outline = skipped_outline_path(task, args.resume)
        if outline:
            mark_progress(progress_path, progress, task, "done", outline=outline)
    if skipped:
        print(f"skipped={skipped}")

    if args.workers == 1:
        for task in runnable:
            mark_progress(progress_path, progress, task, "running")
            result = run_task(
                task,
                args.part_seconds,
                args.resume,
                skip_clean=args.skip_clean,
                skip_context_infer=args.skip_context_infer,
            )
            if result.returncode == 0:
                mark_progress(
                    progress_path,
                    progress,
                    task,
                    "done",
                    outline=result.outline,
                )
                print(f"done={task.progress_key}")
            else:
                error = (result.stderr or result.stdout or "").strip()[-2000:]
                mark_progress(progress_path, progress, task, "error", error=error)
                print(f"error={task.progress_key} returncode={result.returncode}")
        return

    with ProcessPoolExecutor(max_workers=args.workers) as executor:
        future_to_task = {}
        for task in runnable:
            mark_progress(progress_path, progress, task, "running")
            future = executor.submit(
                run_task,
                task,
                args.part_seconds,
                args.resume,
                skip_clean=args.skip_clean,
                skip_context_infer=args.skip_context_infer,
            )
            future_to_task[future] = task

        for future in as_completed(future_to_task):
            task = future_to_task[future]
            try:
                result = future.result()
            except Exception as exc:
                mark_progress(progress_path, progress, task, "error", error=str(exc))
                print(f"error={task.progress_key} {exc}")
                continue

            if result.returncode == 0:
                mark_progress(
                    progress_path,
                    progress,
                    task,
                    "done",
                    outline=result.outline,
                )
                print(f"done={task.progress_key}")
            else:
                error = (result.stderr or result.stdout or "").strip()[-2000:]
                mark_progress(progress_path, progress, task, "error", error=error)
                print(f"error={task.progress_key} returncode={result.returncode}")


if __name__ == "__main__":
    main()
